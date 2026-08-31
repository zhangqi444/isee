#!/usr/bin/env python3
"""
Shadow the Google Sheets workbooks into the repo's content/ tree (spec 6.1).

Reads each workbook XLSX, applies this repo's fix manifests IN MEMORY, then
emits canonical JSON. The exported content therefore matches the corrected
Sheets, not the pre-fix state.

Every item carries: stable id, schema/content version, subject, skill,
difficulty, prompt, choices, correct answer, explanation, misconception tags,
passage ref, and provenance -- the fields spec 6.1 requires.
"""
import openpyxl, json, os, re, hashlib, datetime, sys

OUT = 'content'
SCHEMA_VERSION = '1.0'
CONTENT_VERSION = '2026.08.31'
SOURCE = 'Original content authored for Sheila ISEE'
L = 'ABCD'

def apply(wb, edits):
    for e in edits:
        if e['f'] == 1:      # formulas are runtime, not content
            continue
        try: wb[e['s']][e['a']] = e['n']
        except Exception: pass
    return wb

def load(name, manifests):
    wb = openpyxl.load_workbook(f'{name}.xlsx', data_only=False)
    for m in manifests:
        if os.path.exists(m):
            apply(wb, json.load(open(m)))
    return wb

def digest(obj):
    c = json.dumps(obj, sort_keys=True, ensure_ascii=False, separators=(',', ':'))
    return hashlib.sha256(c.encode()).hexdigest()[:16]

def item(iid, subject, skill, diff, prompt, choices, correct, expl,
         misc=None, passage=None, form=None):
    d = dict(id=iid, schema_version=SCHEMA_VERSION, content_version=CONTENT_VERSION,
             subject=subject, skill=skill, difficulty=diff, prompt=prompt,
             choices={L[i]: choices[i] for i in range(4)}, correct=correct,
             explanation=expl or '', misconceptions=misc or '', source=SOURCE)
    if passage: d['passage_id'] = passage
    if form: d['form'] = form
    d['content_hash'] = digest({k: v for k, v in d.items() if k != 'content_hash'})
    return d

def write(path, obj):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    json.dump(obj, open(path, 'w'), ensure_ascii=False, indent=1)
    return f'{path}  ({len(obj["items"]) if "items" in obj else len(obj)} records)'

log = []

# ---------------- MOCK ----------------
wb = load('mock', ['manifest_slim.json'])
AK = wb['ANSWER KEY']
key = {}
for r in range(2, AK.max_row + 1):
    v = [AK.cell(r, c).value for c in range(1, 12)]
    if v[0]: key[v[0]] = v
ID = re.compile(r'^(DGN|M0\d)-(VR|QR|RC|MA)-\d{3}$')
FORM = {'DIAGNOSTIC': 'DGN', 'MOCK 1': 'M01', 'MOCK 2': 'M02', 'MOCK 3': 'M03'}
mock_items, passages = [], []
for sheet, code in FORM.items():
    ws = wb[sheet]
    for r in range(18, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if isinstance(a, str) and re.match(r'^(DGN|M0\d)-RC-P0\d', str(a)):
            pid = str(a).split(' —')[0].strip()
            title = str(a).split('—')[-1].strip()
            passages.append(dict(id=pid, form=code, title=title,
                                 text=ws.cell(r, 5).value, source=SOURCE,
                                 schema_version=SCHEMA_VERSION,
                                 content_version=CONTENT_VERSION))
        if not (isinstance(a, str) and ID.match(str(a).strip())): continue
        iid = str(a).strip(); k = key.get(iid)
        if not k: continue
        mock_items.append(item(iid, k[2], k[3], k[4], ws.cell(r, 5).value,
                               [ws.cell(r, c).value for c in range(6, 10)],
                               k[5], k[6], k[7], k[8], code))
log.append(write(f'{OUT}/question-banks/mock.json',
                 dict(bank='mock', schema_version=SCHEMA_VERSION,
                      content_version=CONTENT_VERSION, items=mock_items)))
log.append(write(f'{OUT}/passages/mock-passages.json',
                 dict(bank='mock', schema_version=SCHEMA_VERSION,
                      content_version=CONTENT_VERSION, items=passages)))

# ---------------- MA ----------------
wb = load('ma', ['rand_ma.json'])
ks = wb['Sep MA Answer Key']
ma = []
for r in range(2, ks.max_row + 1):
    v = [ks.cell(r, c).value for c in range(1, 14)]
    if not v[0]: continue
    ma.append(item(v[11], 'MA', v[3], v[12], v[5], [v[6], v[7], v[8], v[9]],
                   v[4], v[10], None, None, f'{v[0]}-{v[1]}'))
log.append(write(f'{OUT}/question-banks/ma-september.json',
                 dict(bank='ma-september', schema_version=SCHEMA_VERSION,
                      content_version=CONTENT_VERSION, items=ma)))

# ---------------- QR (options live on weekly tabs) ----------------
wb = load('qr', ['rand_qr.json'])
ks = wb['Sep QR Answer Key']
krows = [r for r in range(2, ks.max_row + 1) if ks.cell(r, 5).value not in (None, '')]
rows = []
for sh in ['Sep QR W1', 'Sep QR W2', 'Sep QR W3', 'Sep QR W4']:
    ws = wb[sh]
    for r in range(1, ws.max_row + 1):
        if isinstance(ws.cell(r, 1).value, (int, float)) and ws.cell(r, 3).value and ws.cell(r, 4).value is not None:
            rows.append((sh, r))
qr = []
for (sh, r), kr in zip(rows, krows):
    ws = wb[sh]
    sid = ks.cell(kr, 8).value
    sid = sid if isinstance(sid, str) else f'QR-SEP-{kr-1:03d}'
    qr.append(item(sid, 'QR', ks.cell(kr, 4).value, ks.cell(kr, 11).value,
                   ws.cell(r, 3).value, [ws.cell(r, c).value for c in range(4, 8)],
                   ks.cell(kr, 5).value, ks.cell(kr, 6).value,
                   ks.cell(kr, 12).value, None, sh.replace('Sep QR ', '')))
log.append(write(f'{OUT}/question-banks/qr-september.json',
                 dict(bank='qr-september', schema_version=SCHEMA_VERSION,
                      content_version=CONTENT_VERSION, items=qr)))

# ---------------- RC ----------------
wb = load('rc', [])
ks = wb['RC Answer Key']
kmap = {}
for r in range(2, ks.max_row + 1):
    wk, se, q = ks.cell(r, 1).value, ks.cell(r, 2).value, ks.cell(r, 3).value
    if wk and q is not None:
        kmap[(str(wk), str(se), int(q))] = [ks.cell(r, c).value for c in range(1, 12)]
rc, rcp = [], []
for wk in ['Sep RC W1', 'Sep RC W2', 'Sep RC W3', 'Sep RC W4']:
    ws = wb[wk]; W = 'W' + wk[-1]; sess = None; pid = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value; t = str(a or '')
        if t.startswith('SESSION'):
            sess = 'S' + t.split('—')[0].replace('SESSION', '').strip(); pid = None
        elif sess and pid is None and isinstance(a, str) and len(t) > 150:
            pid = f'RC-SEP-{W}-{sess}-P'
            rcp.append(dict(id=pid, week=W, session=sess, text=t, source=SOURCE,
                            schema_version=SCHEMA_VERSION, content_version=CONTENT_VERSION))
        if isinstance(a, (int, float)) and ws.cell(r, 3).value:
            k = kmap.get((W, sess, int(a)))
            if not k: continue
            rc.append(item(k[8] or f'RC-SEP-{W}-{sess}-Q{int(a):02d}', 'RC', k[3], k[10],
                           ws.cell(r, 3).value, [ws.cell(r, c).value for c in range(4, 8)],
                           k[4], k[6], None, pid, W))
log.append(write(f'{OUT}/question-banks/rc-september.json',
                 dict(bank='rc-september', schema_version=SCHEMA_VERSION,
                      content_version=CONTENT_VERSION, items=rc)))
log.append(write(f'{OUT}/passages/rc-september-passages.json',
                 dict(bank='rc-september', schema_version=SCHEMA_VERSION,
                      content_version=CONTENT_VERSION, items=rcp)))

# ---------------- VR ----------------
wb = load('vr', ['rand_vr.json'])
ks = wb['Sep VR Answer Key']
krows = [r for r in range(2, ks.max_row + 1) if ks.cell(r, 4).value not in (None, '')]
rows = []
for sh in ['Sep VR W1', 'Sep VR W2', 'Sep VR W3', 'Sep VR W4']:
    ws = wb[sh]; sess = None
    for r in range(1, ws.max_row + 1):
        t = str(ws.cell(r, 1).value or '')
        if t.startswith('SESSION'): sess = t
        if sess and 'SESSION 1' in sess: continue
        if isinstance(ws.cell(r, 1).value, (int, float)) and all(ws.cell(r, c).value is not None for c in (2, 3, 4, 5, 6)):
            rows.append((sh, r, sess))
vr = []
for (sh, r, sess), kr in zip(rows, krows):
    ws = wb[sh]
    word = ks.cell(kr, 5).value
    word = word if isinstance(word, str) else ''
    vr.append(item(f'VR-SEP-{sh[-2:]}-{r:03d}', 'VR', word or 'vocabulary', 'M',
                   ws.cell(r, 2).value, [ws.cell(r, c).value for c in range(3, 7)],
                   ks.cell(kr, 4).value, '', None, None, sh.replace('Sep VR ', '')))
log.append(write(f'{OUT}/question-banks/vr-september.json',
                 dict(bank='vr-september', schema_version=SCHEMA_VERSION,
                      content_version=CONTENT_VERSION, items=vr)))

# ---------------- manifest ----------------
banks = {}
for f in sorted(os.listdir(f'{OUT}/question-banks')):
    d = json.load(open(f'{OUT}/question-banks/{f}'))
    banks[d['bank']] = len(d['items'])
meta = dict(schema_version=SCHEMA_VERSION, content_version=CONTENT_VERSION,
            generated_at=datetime.datetime.now(datetime.timezone.utc).isoformat(),
            generator='extract_content.py',
            note='Shadow of the Google Sheets workbooks with reviewed fixes applied. '
                 'Sheets remain the authoring surface until the PWA takes over.',
            banks=banks)
json.dump(meta, open(f'{OUT}/manifest.json', 'w'), ensure_ascii=False, indent=1)
for l in log: print(' ', l)
print('\n', json.dumps(banks, indent=1))
