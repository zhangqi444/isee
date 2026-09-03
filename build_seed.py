#!/usr/bin/env python3
"""Migrate Sheila's completed Week-1 work from the Sheets into the site's
progress model. Reads her per-item answers, matches each to the site bundle by
question text, and emits a seed keyed exactly like the app's own results
(setId -> {n,right,at,wrong[]}). The app applies it once, per viewer, without
clobbering any work she later does on the site itself."""
import openpyxl, json, re

def V(ws,r,c):
    v=ws.cell(r,c).value
    if v is None: return None
    if isinstance(v,str):
        s=v.strip(); return None if (s=='' or s.startswith('=')) else s
    return v
def norm(s):
    return re.sub(r'\s+',' ',re.sub(r'[^a-z0-9 ]',' ',str(s).lower())).strip()

bundle=json.load(open('site/content/bundle.json'))
SETSIZE=12
DATE='2026-08-31T12:00:00Z'  # her actual W1 submission window

# (subject-key, xlsx, sheet, [(answer_col, correct_col, [rows...])], question_col)
SPECS = {
 'ma': ('ma_cur.xlsx','Sep MA W1', 8,10, list(range(7,19))+list(range(24,36)), 3),
 'vr': ('vr_check.xlsx','Sep VR W1', 7,10, list(range(30,50))+list(range(54,71)), 2),
 'qr': ('qr_cur.xlsx','Sep QR W1', 8,10, list(range(7,19))+list(range(24,39)), 3),
 'rc': ('rc_cur.xlsx','Sep RC W1', 8,11, list(range(7,13))+list(range(20,26)), 3),
}

results={}; report=[]
for sub,(xlsx,sheet,acol,ccol,rows,qcol) in SPECS.items():
    ws=openpyxl.load_workbook(xlsx,data_only=True)[sheet]
    # her verdict per normalized question
    verdict={}
    for r in rows:
        a=V(ws,r,acol)
        if not a: continue
        c=V(ws,r,ccol)
        q=norm(ws.cell(r,qcol).value)
        verdict[q]=(str(a).strip()==str(c).strip(), str(a).strip().upper()[:1])
    # site's W1 items for this subject, in the app's own order, sliced into sets
    items=[it for it in bundle['subjects'][sub] if it['w']=='W1']
    k=-(-len(items)//SETSIZE); base,extra=divmod(len(items),k)   # balanced, mirrors content.js chunk()
    sets=[]; at=0
    for i in range(k):
        n=base+(1 if i<extra else 0); sets.append(items[at:at+n]); at+=n
    matched=unmatched=0
    for n,st in enumerate(sets):
        right=0; wrong=[]; seen=0; picks={}
        for it in st:
            v=verdict.get(norm(it['q']))
            if v is None: continue        # she didn't answer this site item
            seen+=1
            ok,letter=v
            if letter in 'ABCD': picks[it['id']]=letter
            if ok: right+=1
            else: wrong.append(it['id'])
        if seen:
            results[f'{sub}:W1:{n}']={'n':len(st),'right':right,'at':DATE,'wrong':wrong,'picks':picks}
        matched+=seen; unmatched+=len(st)-seen
    report.append((sub,len(items),len(sets),matched,unmatched,
                   sum(results[k]['right'] for k in results if k.startswith(sub)),
                   sum(len(results[k]['wrong']) for k in results if k.startswith(sub))))

seed={'version':3,'migrated_at':DATE,
      'note':"Sheila's Week-1 work, migrated from the Google Sheets on 2026-09-01 (v3: balanced set sizes + her chosen letters).",
      'results':results}
json.dump(seed,open('site/content/seed.json','w'),ensure_ascii=False,indent=1)

print(f"{'sub':4} {'items':>5} {'sets':>4} {'matched':>7} {'unmatched':>9} {'right':>5} {'wrong':>5}")
for row in report: print(f"{row[0]:4} {row[1]:5d} {row[2]:4d} {row[3]:7d} {row[4]:9d} {row[5]:5d} {row[6]:5d}")
tot_r=sum(r[5] for r in report); tot_w=sum(r[6] for r in report)
print(f"\nTOTAL migrated: {tot_r} correct, {tot_w} to review, across {len(results)} sets")
print("setIds:", ', '.join(sorted(results)))
